"""
noi_soi_parser.py — Parse Nhật Ký Nối Sợi + Tạo output 3-sheet merged
"""
import re, io, os, copy
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

BLUE_COLOR = "FF00B0F0"


# ─── 1. TIME PARSING ────────────────────────────────────────────────────────
def _parse_minutes(s: str) -> int:
    """✅ FIX: Handles chained ranges '19H-21H+22H-6H' = 120+480 = 600 (not 120+22=142)."""
    if not s or str(s).strip() in ('…', '', 'None', 'nan', '...'): return 0
    token = str(s).strip()
    m = re.search(r'(\d+)h(\d*)[-→>–—]+(\d+)h(\d*)', token, re.IGNORECASE)
    if not m:
        # ✅ FIX: Nếu token trông như "XhYY-..." (có giờ bắt đầu nhưng ko có giờ kết thúc)
        # → thời gian không xác định → trả 0 thay vì lấy nhầm số giờ/phút
        _has_start  = re.search(r'\d+h\d*', token, re.IGNORECASE)
        _has_dash   = re.search(r'[-→>–—]', token)
        _has_end_h  = re.search(r'[-→>–—]\s*\d+h', token, re.IGNORECASE)
        if _has_start and _has_dash and not _has_end_h:
            return 0  # incomplete range like "15h35-..." or "23h-"
        # No time range — try plain addition '60+30' or single number
        if '+' in token:
            nums = [int(n) for n in re.findall(r'\d+', token) if 1 <= int(n) < 600]
            return sum(nums) if nums else 0
        nums = re.findall(r'\d+', token)
        valid = [int(n) for n in nums if 1 <= int(n) < 600]
        return valid[-1] if valid else 0
    h1, m1 = int(m.group(1)), int(m.group(2) or 0)
    h2, m2 = int(m.group(3)), int(m.group(4) or 0)
    diff = (h2*60+m2) - (h1*60+m1)
    if diff < 0: diff += 1440
    rest = token[m.end():]
    extras = 0
    # ✅ FIX: Look for additional time RANGES after + (e.g. '+22H-6H' = 480 min)
    for er in re.finditer(r'[+]\s*(\d+)h(\d*)[-→>–—]+(\d+)h(\d*)', rest, re.IGNORECASE):
        eh1, em1 = int(er.group(1)), int(er.group(2) or 0)
        eh2, em2 = int(er.group(3)), int(er.group(4) or 0)
        ed = (eh2*60+em2) - (eh1*60+em1)
        extras += ed if ed >= 0 else ed + 1440
    # Plain +N additions (only where no time range was found)
    plain = re.sub(r'[+]\s*\d+h\d*[-→>–—]+\d+h\d*', '', rest, flags=re.IGNORECASE)
    extras += sum(int(n) for n in re.findall(r'(?<=[+])\s*(\d+)', plain) if 1 <= int(n) < 600)
    return diff + extras


# ─── 2. FILE TYPE DETECTION ─────────────────────────────────────────────────
def _detect_weaving(filename: str) -> str:
    """'1동' or 'wea 1' → 'Weaving 1' | '2동'→'Weaving 2' | '3동'→'Weaving 3'"""
    fn = filename.lower()
    if '1동' in fn or 'wea 1' in fn or 'wea1' in fn or '_1_' in fn or '제직_1' in fn: return 'Weaving 1'
    if '2동' in fn or 'wea 2' in fn or 'wea2' in fn or '_2_' in fn or '제직_2' in fn: return 'Weaving 2'
    if '3동' in fn or 'wea 3' in fn or 'wea3' in fn or '_3_' in fn or '제직_3' in fn: return 'Weaving 3'
    return 'UNKNOWN'

def _is_timing_file(filename: str) -> bool:
    fn = filename.lower()
    return any(k in fn for k in ['nhật ký', 'nhat ky', '타이밍', 'noi soi', 'timing'])

def _timing_file_weaving(filename: str) -> str:
    fn = filename.lower()
    if 'wea 2' in fn or 'wea2' in fn or '_wea_2' in fn: return 'W2'
    return 'W1+3'


# ─── 3. SHEET FINDER ────────────────────────────────────────────────────────
def _find_sheet(wb: openpyxl.Workbook, target_date: str) -> str | None:
    if not target_date: return wb.sheetnames[0] if wb.sheetnames else None
    m = re.search(r'(\d{1,2})[/-](\d{1,2})', str(target_date))
    if not m: return wb.sheetnames[0] if wb.sheetnames else None
    day, mon = int(m.group(1)), int(m.group(2))
    for name in wb.sheetnames:
        nm = re.match(r'^(\d{1,2})[.-](\d{1,2})$', name.strip())
        if nm and int(nm.group(1)) == day and int(nm.group(2)) == mon:
            return name
    return wb.sheetnames[0] if wb.sheetnames else None

def get_available_sheets(file_obj, filename: str = '') -> list:
    fn = filename or getattr(file_obj, 'name', '')
    try:
        data = file_obj.read() if hasattr(file_obj, 'read') else open(file_obj, 'rb').read()
        if hasattr(file_obj, 'seek'): file_obj.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        names = wb.sheetnames; wb.close(); return names
    except: return []


# ─── 4. PARSE NHẬT KÝ SHEET ─────────────────────────────────────────────────
def _parse_sheet(ws, file_weaving: str) -> list:
    agg = {}
    for row in ws.iter_rows(min_row=6):
        cell_a = row[0] if row else None
        if not cell_a or cell_a.value is None: continue
        val_str = str(cell_a.value).strip()
        if re.match(r'^CA\s+[ABC]', val_str, re.IGNORECASE): continue
        if not re.match(r'^[\d.]+$', val_str.replace(' ','')): continue

        # Blue cell or .3 suffix → Weaving 3
        is_blue = False
        try:
            fill = cell_a.fill
            if fill and fill.fgColor and fill.fgColor.type == 'rgb':
                is_blue = fill.fgColor.rgb == BLUE_COLOR
        except: pass

        has_dot3 = '.3' in val_str
        if is_blue or has_dot3:
            weaving = 'Weaving 3'; mac_num = int(float(val_str.replace('.3','').strip() or 0))
        elif file_weaving == 'W2':
            weaving = 'Weaving 2'; mac_num = int(float(val_str))
        else:
            weaving = 'Weaving 1'; mac_num = int(float(val_str))
        if mac_num <= 0: continue

        is_tren = str(row[1].value if len(row) > 1 else '').strip() == '+'
        is_duoi = str(row[2].value if len(row) > 2 else '').strip() == '+'
        t_noi = _parse_minutes(str(row[7].value if len(row) > 7 else ''))
        t_keo = _parse_minutes(str(row[8].value if len(row) > 8 else ''))
        ten_hang = str(row[4].value if len(row) > 4 else '').strip()
        loai_soi = str(row[3].value if len(row) > 3 else '').strip()
        ma_beam  = str(row[11].value if len(row) > 11 else '').strip()

        key = (weaving, mac_num)
        if key not in agg:
            agg[key] = {'weaving': weaving, 'so_may': mac_num,
                        'tyming_tren':0,'tyming_duoi':0,'pull_beam_tren':0,'pull_beam_duoi':0,
                        'ten_hang': ten_hang, 'loai_soi': loai_soi, 'ma_beam': ma_beam}
        if is_tren:
            agg[key]['tyming_tren']    += t_noi
            agg[key]['pull_beam_tren'] += t_keo
        elif is_duoi:
            agg[key]['tyming_duoi']    += t_noi
            agg[key]['pull_beam_duoi'] += t_keo
        else:
            agg[key]['tyming_tren'] += t_noi // 2
            agg[key]['tyming_duoi'] += t_noi // 2
        if ten_hang and not agg[key]['ten_hang']: agg[key]['ten_hang'] = ten_hang
    return list(agg.values())

def parse_file(file_obj, target_date: str = None, filename: str = '') -> dict:
    fn = filename or getattr(file_obj, 'name', '')
    fw = _timing_file_weaving(fn)
    data = file_obj.read() if hasattr(file_obj, 'read') else open(file_obj,'rb').read()
    if hasattr(file_obj, 'seek'): file_obj.seek(0)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_name = _find_sheet(wb, target_date)
    if not sheet_name: return {'error': 'Không tìm thấy sheet', 'records': []}
    ws = wb[sheet_name]
    date_val = ''
    try:
        v = ws.cell(2,5).value
        date_val = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else str(v)[:10]
    except: date_val = target_date or ''
    records = _parse_sheet(ws, fw)
    wb.close()
    return {'filename': fn, 'file_weaving': fw, 'sheet_used': sheet_name,
            'date': date_val, 'records': records}


# ─── 5. DETECT TIMING COLUMNS IN ANALYSIS SHEET ──────────────────────────────
def _detect_timing_cols(ws) -> dict:
    """
    Scan row 5 for UP(Trên)/DOWN(Dưới) headers.
    Returns {tyming_tren, tyming_duoi, pull_beam_tren, pull_beam_duoi} as 1-indexed cols.
    Falls back to known positions if not found.
    """
    up_cols = []; dn_cols = []
    header_row = list(ws.iter_rows(min_row=5, max_row=5))[0]
    for cell in header_row:
        v = str(cell.value or '').replace('\n', ' ')
        if 'UP' in v.upper() and 'Trên' in v: up_cols.append(cell.column)
        elif 'DOWN' in v.upper() and 'Dưới' in v: dn_cols.append(cell.column)

    if len(up_cols) >= 2 and len(dn_cols) >= 2:
        return {'tyming_tren': up_cols[0], 'tyming_duoi': dn_cols[0],
                'pull_beam_tren': up_cols[1], 'pull_beam_duoi': dn_cols[1]}

    # Fallback: detect from row 4 (TYMING header)
    for cell in list(ws.iter_rows(min_row=4, max_row=4))[0]:
        v = str(cell.value or '')
        if 'TYMING' in v.upper() and 'NỐI' in v.upper():
            base = cell.column
            return {'tyming_tren': base, 'tyming_duoi': base+1,
                    'pull_beam_tren': base+2, 'pull_beam_duoi': base+3}

    # Last fallback by max_column heuristic
    max_col = ws.max_column
    return {'tyming_tren': 18 if max_col > 50 else 12,
            'tyming_duoi': 19 if max_col > 50 else 13,
            'pull_beam_tren': 20 if max_col > 50 else 14,
            'pull_beam_duoi': 21 if max_col > 50 else 15}


# ─── 6. COPY SHEET WITH FORMATTING ──────────────────────────────────────────
def _copy_sheet(source_ws, target_wb: Workbook, new_name: str):
    """Deep-copy source_ws into target_wb with formatting preserved."""
    target_ws = target_wb.create_sheet(title=new_name[:31])

    # Copy cells with styles
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column)
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font        = copy.copy(cell.font)
                new_cell.border      = copy.copy(cell.border)
                new_cell.fill        = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection  = copy.copy(cell.protection)
                new_cell.alignment   = copy.copy(cell.alignment)

    # Merged cells
    for rng in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(rng))

    # Row heights
    for row_num, rd in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = rd.height

    # Column widths
    for col_ltr, cd in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_ltr].width = cd.width

    # Tab color (optional)
    try: target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor
    except: pass

    return target_ws


# ─── 7. FILL TIMING INTO SHEET ───────────────────────────────────────────────
def _fill_timing(ws, timing_records: list, weaving_filter: str):
    """Fill tyming_tren/duoi/pull_beam_tren/duoi into ws rows matched by so_may."""
    cols = _detect_timing_cols(ws)
    timing_map = {}
    for r in timing_records:
        if weaving_filter and r['weaving'] != weaving_filter: continue
        mac = int(r['so_may'])
        if mac not in timing_map:
            timing_map[mac] = {k: 0 for k in cols}
        for field, col in cols.items():
            timing_map[mac][field] = timing_map[mac].get(field, 0) + r.get(field, 0)

    filled = 0
    for row_idx in range(6, ws.max_row + 1):
        v = ws.cell(row_idx, 1).value
        if v is None: continue
        try: mac = int(float(str(v).replace('.3','').strip()))
        except: continue
        if mac not in timing_map: continue
        for field, col in cols.items():
            val = timing_map[mac].get(field, 0)
            if val > 0:
                existing = ws.cell(row_idx, col).value
                if not existing or existing == 0:
                    ws.cell(row_idx, col).value = val
                    filled += 1
    return filled


# ─── 8. MAIN: CREATE MERGED 3-SHEET OUTPUT ──────────────────────────────────
def create_merged_output(files_info: list, timing_records: list, date_str: str = '') -> bytes:
    """
    files_info: [{'bytes': bytes, 'filename': str, 'weaving': 'Weaving 1'|2|3}, ...]
    Creates ONE Excel with 3 sheets (one per weaving), named 'Weaving X (date)'.
    Fills BOTH:
      - Downtime columns (from parsing notes in each analysis sheet)
      - Timing columns   (from nhật ký nối sợi records)
    """
    out_wb = Workbook()
    if out_wb.active.title == 'Sheet':
        out_wb.remove(out_wb.active)

    date_label   = f" ({date_str})" if date_str else ''
    filled_total = 0

    for info in sorted(files_info, key=lambda x: x['weaving']):
        src_wb = openpyxl.load_workbook(io.BytesIO(info['bytes']), data_only=True)
        if 'analysis' not in src_wb.sheetnames:
            src_wb.close(); continue

        src_ws     = src_wb['analysis']
        sheet_name = f"{info['weaving']}{date_label}"
        target_ws  = _copy_sheet(src_ws, out_wb, sheet_name)
        src_wb.close()

        # ── Step 1: Fill downtime columns from parsed notes ──
        note_map = _read_note_map(target_ws)
        n1 = _fill_all_downtime(target_ws, note_map)

        # ── Step 2: Fill timing columns from nhật ký ──
        n2 = _fill_timing(target_ws, timing_records, info['weaving'])

        filled_total += n1 + n2

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue(), filled_total


# ─── 9. LEGACY: single-file merge ────────────────────────────────────────────
def merge_into_analysis(analysis_bytes: bytes, timing_records: list,
                        weaving_filter: str = None) -> tuple:
    wb = openpyxl.load_workbook(io.BytesIO(analysis_bytes))
    if 'analysis' not in wb.sheetnames: return analysis_bytes, 0
    ws = wb['analysis']
    n = _fill_timing(ws, timing_records, weaving_filter)
    out = io.BytesIO(); wb.save(out)
    return out.getvalue(), n


# ─── 10. NOTE READING + DOWNTIME FILLING ─────────────────────────────────────
def _nfc(s):
    """Normalize Unicode to NFC so 'máy' matches both composed and decomposed forms."""
    import unicodedata as _ud
    return _ud.normalize('NFC', str(s))


def _detect_note_structure(ws):
    """
    Returns (note_col, use_label):
    - W1/W2/W3 style: 'máy N' label at col L, note at col L+2 → (L+2, True)
    - Fallback: note at col 45, same row as so_may → (45, False)
    ✅ FIX: normalize Unicode NFC before matching so decomposed 'máy' (a+U+0301) works.
    """
    for row in ws.iter_rows(min_row=6, max_row=min(40, ws.max_row)):
        for cell in row:
            if cell.value and re.match(r'^máy\s*\d+$', _nfc(str(cell.value)).strip(), re.IGNORECASE):
                return cell.column + 2, True
    # Check col 45 for text notes
    tc = sum(1 for r in range(6, min(20, ws.max_row+1))
             if ws.cell(r, 45).value and len(str(ws.cell(r,45).value or '').strip()) > 5)
    return (45, False) if tc > 0 else (None, False)


def _read_note_map(ws) -> dict:
    """Returns {so_may_int: note_text} for all machines."""
    note_col, use_label = _detect_note_structure(ws)
    if not note_col:
        return {}
    note_map = {}
    if use_label:
        label_col = note_col - 2
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            lv = _nfc(str(row[label_col-1].value or '')).strip()  # ✅ NFC normalize
            m  = re.match(r'máy\s*(\d+)', lv, re.IGNORECASE)
            if m:
                mac = int(m.group(1))
                nv  = str(row[note_col-1].value or '').strip()
                if nv and len(nv) > 1 and not nv.replace('.','').replace('-','').isdigit():
                    note_map[mac] = nv
    else:
        for row_idx in range(6, ws.max_row + 1):
            sv = ws.cell(row_idx, 1).value
            if sv is None: continue
            try: mac = int(float(str(sv).replace('.3','').strip()))
            except: continue
            nv = str(ws.cell(row_idx, note_col).value or '').strip()
            if nv and len(nv) > 3 and not nv.replace('.','').replace('-','').isdigit():
                note_map[mac] = nv
    return note_map


def _detect_col_offset(ws) -> int:
    """Physical openpyxl col = COL_IDX_key + offset. Detect by finding 'no order' in row 5."""
    for cell in ws[5]:
        if 'no order' in str(cell.value or '').lower().replace('\n',' '):
            return cell.column - 1  # COL_IDX[1]=no_order → physical = 1 + offset
    return 7  # default W1/W2


def _fill_all_downtime(ws, note_map: dict):
    """
    Parse each machine's note → fill ALL downtime columns (COL_IDX 1-10).
    Uses analysis_parser.parse_note if available, else inline.
    """
    if not note_map:
        return 0

    # Try import analysis_parser
    try:
        from analysis_parser import parse_note as _pn
        _parse = _pn
    except ImportError:
        _parse = None

    if _parse is None:
        # Inline minimal parser (subset of KEYWORD_MAP)
        def _parse(note: str) -> dict:
            """
            ✅ FIX: Uses _parse_minutes (fixed chained +HH-HH ranges).
            ✅ FIX: Extended INLINE_MAP covering all keywords including after /.
            """
            import re as _re
            result = {}
            # Ordered by specificity (more specific first to avoid false matches)
            INLINE_MAP = [
                ('mc stop',           'mc_stop'),
                ('đóng máy',          'mc_stop'),   ('dong may',    'mc_stop'),
                ('rối',               'beam_roi'),   ('roi ',        'beam_roi'),
                ('rói',               'beam_roi'),
                ('go + lỗi',          'loi'),
                ('lỗi go',            'loi_go_jq'), ('loi go',      'loi_go_jq'),
                ('jq',                'loi_go_jq'),
                ('go ',               'loi_go_jq'),
                ('kiếm đai',          'kiem_dai'),  ('kiem dai',    'kiem_dai'),
                ('ktkm',              'kiem_dai'),
                ('đứt sợi ngang',     'changed_broke'), ('dut soi ngang', 'changed_broke'),
                ('hư sợi ngang',      'changed_broke'), ('hu soi ngang',  'changed_broke'),
                ('đứt sợi',           'changed_broke'), ('dut soi',       'changed_broke'),
                ('đứt g',             'changed_broke'), ('đứt p',         'changed_broke'),
                ('dsp',               'changed_broke'), ('đsp',           'changed_broke'),
                ('hbp+g',             'sua_may_khac'), ('hbg+p',         'sua_may_khac'),
                ('hbg',               'sua_may_khac'), ('hbp',           'sua_may_khac'),
                ('dao cắt',           'sua_may_khac'), ('dao cat',       'sua_may_khac'),
                ('lăn gai',           'sua_may_khac'), ('lan gai',       'sua_may_khac'),
                ('sọc viền',          'sua_may_khac'), ('soc vien',      'sua_may_khac'),
                ('rách khăn',         'sua_may_khac'), ('rach khan',     'sua_may_khac'),
                ('haness',            'sua_may_khac'),
                ('thay dây curoa',    'sua_may_khac'), ('thay day curoa','sua_may_khac'),
                ('cắt beam',          'sua_may_khac'), ('cat beam',      'sua_may_khac'),
                ('thay beam',         'pull_beam_tren'),
                ('xử lí sợi đứt',     'sua_may_khac'), ('xu li soi',    'sua_may_khac'),
                ('móc p',             'sua_may_khac'), ('moc p',         'sua_may_khac'),
                ('vệ sinh',           'clean_mc'),     ('ve sinh',       'clean_mc'),
                ('sửa máy',           'sua_may_khac'), ('sua may',       'sua_may_khac'),
                ('lỗi',               'loi'),          ('loi ',          'loi'),
            ]
            for part in _re.split(r'[,/]', str(note)):
                part = part.strip()
                if not part: continue
                # ✅ FIX: call module _parse_minutes (handles chained +HH-HH ranges)
                mins = _parse_minutes(part)
                if not mins: continue
                pl = _re.sub(r'\s+', ' ', part.lower().strip())
                matched = False
                for kw, field in INLINE_MAP:
                    if kw.lower() in pl:
                        result[field] = result.get(field, 0) + mins
                        matched = True
                        break
                if not matched and any(c.isalpha() for c in part):
                    # Unknown keyword with time → sua_may_khac as safe fallback
                    result['sua_may_khac'] = result.get('sua_may_khac', 0) + mins
            return result

    offset = _detect_col_offset(ws)

    # COL_IDX field → openpyxl column
    FIELD_TO_IDX = {
        'no_order':1,'mc_stop':2,'thay_go':3,'loi_go_jq':4,'ccsn':5,
        'brake':6,'cutter':7,'loi':8,'kiem_dai':9,'sua_may_khac':10,
        'tyming_tren':11,'tyming_duoi':12,'pull_beam_tren':13,'pull_beam_duoi':14,
        'cho':15,'no_beam_wea':16,'beam_roi':20,'clean_mc':22,'sample_sample':28,
        'changed_pm':31,'changed_broke':32,'mc_stop':2,
    }

    filled = 0
    for row_idx in range(6, ws.max_row + 1):
        sv = ws.cell(row_idx, 1).value
        if sv is None: continue
        try: mac = int(float(str(sv).replace('.3','').strip()))
        except: continue
        note = note_map.get(mac, '')
        if not note or note == '0': continue
        parsed = _parse(note)
        for field, mins in parsed.items():
            if mins <= 0: continue
            cidx = FIELD_TO_IDX.get(field)
            if cidx is None: continue
            phys_col = cidx + offset
            if phys_col < 1 or phys_col > ws.max_column: continue
            existing = ws.cell(row_idx, phys_col).value
            if not existing or existing == 0:
                ws.cell(row_idx, phys_col).value = mins
                filled += 1
    return filled